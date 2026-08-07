"""Validate the clean Python model against the cached values in the workbook.

Compares the mite trajectory (start and end of each period) and the alcohol-wash
count against the "Current version" sheet of
"Randys-Varroa-Model-V2026 Web(1).xlsx".

The workbook's default run uses:
  colony type "d", starting mites 100, immigration setting 0 (no neighbours),
  one treatment: 80% kill in period 16 (cell S21 = 0.8).

Usage:
    python3 validate.py "path/to/Randys-Varroa-Model-V2026 Web(1).xlsx"
"""

import sys
from typing import TypedDict, cast

import openpyxl

from varroa import VarroaModel

WORKBOOK = sys.argv[1] if len(sys.argv) > 1 else \
    "/home/faraaz/Downloads/Randys-Varroa-Model-V2026 Web(1).xlsx"

TOL = 1e-6


class ReferenceRow(TypedDict):
    """Cached per-period values read from the workbook's 'Current version' sheet."""

    start: float | None
    end: float | None
    r: float | None
    wash: float | None
    invaded: float | None
    phoretic: float | None


def load_reference() -> dict[int, ReferenceRow]:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Current version"]
    ref: dict[int, ReferenceRow] = {}
    for row in range(5, 29):  # periods 1..24
        period = row - 4
        ref[period] = {
            "start": cast(float | None, ws[f"DF{row}"].value),     # mites at start of period
            "end": cast(float | None, ws[f"DS{row}"].value),      # mites at end of period
            "r": cast(float | None, ws[f"DT{row}"].value),        # net r observed for the period
            "wash": cast(float | None, ws[f"DX{row}"].value),     # alcohol-wash count
            "invaded": cast(float | None, ws[f"CU{row}"].value),  # % worker cells invaded
            "phoretic": cast(float | None, ws[f"CA{row}"].value), # % mites phoretic
        }
    return ref


def main():
    ref = load_reference()
    trt = [0.0] * 24
    trt[15] = 0.8  # period 16, cell S21 in the workbook
    run = VarroaModel(colony_type="d", initial_mites=100.0,
                      immigration_setting=0, treatment_kills=trt).run()

    print(f"{'p':>2} {'model_start':>12} {'ref_start':>12} {'model_end':>12} "
          + f"{'ref_end':>12} {'model_wash':>10} {'ref_wash':>10}  match")
    max_err = 0.0
    all_ok = True
    for pr in run.periods:
        r = ref[pr.period]
        ref_start = r["start"] or 0.0
        ref_end = r["end"] or 0.0
        ref_wash = r["wash"] or 0.0
        err = max(abs(pr.mites_start - ref_start), abs(pr.mites_end - ref_end),
                  abs(pr.wash_count - ref_wash))
        max_err = max(max_err, err)
        ok = err < TOL
        all_ok &= ok
        print(f"{pr.period:>2} {pr.mites_start:>12.6f} {ref_start:>12.6f} "
              + f"{pr.mites_end:>12.6f} {ref_end:>12.6f} "
              + f"{pr.wash_count:>10.4f} {ref_wash:>10.4f}  {'OK' if ok else 'MISMATCH'}")

    print(f"\nMax absolute error vs workbook: {max_err:.3e}")
    print("RESULT:", "ALL PERIODS MATCH" if all_ok else "MISMATCHES FOUND")


if __name__ == "__main__":
    main()
